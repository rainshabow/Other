#!/usr/bin/env python3
"""比较 output(result).xlsx 与 input.xlsx 的第一列（或指定列），将 input 中不同的单元格更新为 output 的值并填充黄色。

用法示例：
    python sync_firstcol.py
    等同于：
    python sync_firstcol.py -i input.xlsx -r result.xlsx -c 1
"""
from pathlib import Path
import argparse
import shutil
import time
from typing import Optional

from openpyxl import load_workbook
from openpyxl.styles import PatternFill


YELLOW_FILL = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")


def normalize(v: Optional[object]) -> str:
    if v is None:
        return ""
    return str(v).strip()


def main():
    parser = argparse.ArgumentParser(description="把 input 表的指定列更新为 result 表相应列的内容，并对改变的单元格填充黄色。")
    script_dir = Path(__file__).parent
    parser.add_argument("--input", "-i", default=str(script_dir / "input.xlsx"), help="要修改的输入 Excel（默认脚本同级的 input.xlsx）")
    parser.add_argument("--result", "-r", default=str(script_dir / "result.xlsx"), help="结果 Excel，来源要与之对比（默认脚本同级的 result.xlsx）")
    parser.add_argument("--sheet", "-s", default=None, help="工作表名称或索引，默认使用活动表")
    parser.add_argument("--col", "-c", type=int, default=1, help="要比较和修改的列（1-based，默认 1）")
    parser.add_argument("--backup", action="store_true", help="在修改前创建输入文件的备份")
    parser.add_argument("--dry-run", action="store_true", help="仅打印将要修改的单元格，不实际保存文件")
    args = parser.parse_args()

    input_path = Path(args.input)
    result_path = Path(args.result)

    if not input_path.exists():
        print(f"输入文件不存在: {input_path}")
        return
    if not result_path.exists():
        print(f"结果文件不存在: {result_path}")
        return

    print(f"读取输入文件: {input_path}")
    wb_in = load_workbook(input_path)
    if args.sheet is None:
        ws_in = wb_in.active
    else:
        ws_in = wb_in[args.sheet]

    print(f"读取结果文件: {result_path}")
    wb_res = load_workbook(result_path, read_only=True)
    if args.sheet is None:
        ws_res = wb_res.active
    else:
        ws_res = wb_res[args.sheet]

    col_idx = args.col
    max_row = max(ws_in.max_row, ws_res.max_row)

    if args.backup and not args.dry_run:
        stamp = time.strftime("%Y%m%d_%H%M%S")
        backup_path = input_path.with_name(f"{input_path.stem}.backup.{stamp}{input_path.suffix}")
        shutil.copy(input_path, backup_path)
        print(f"已创建备份: {backup_path}")

    changes = []
    for r in range(1, max_row + 1):
        val_in = normalize(ws_in.cell(row=r, column=col_idx).value)
        val_res = normalize(ws_res.cell(row=r, column=col_idx).value)
        if val_in != val_res:
            changes.append((r, val_in, val_res))

    if not changes:
        print("未发现需要更新的单元格。")
        return

    print(f"发现 {len(changes)} 处差异。")
    if args.dry_run:
        for r, old, new in changes:
            print(f"行 {r}: '{old}' -> '{new}'")
        return

    for r, old, new in changes:
        cell = ws_in.cell(row=r, column=col_idx)
        cell.value = new
        cell.fill = YELLOW_FILL

    try:
        wb_in.save(input_path)
        print(f"已将 {len(changes)} 个单元格更新并保存到: {input_path}")
    except Exception as e:
        print(f"保存失败: {e}")


if __name__ == "__main__":
    main()
